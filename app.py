"""
M&A Analyst Agent — Streamlit Web Interface
Same agent logic as ma_agent.py, runs in the browser.

Local:          streamlit run app.py
Streamlit Cloud: set ANTHROPIC_API_KEY in Secrets
"""

import os
import io
import json
import time
import threading
import traceback
from pathlib import Path
from queue import Queue, Empty

import httpx
import anthropic
import openpyxl
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="M&A Analyst Agent",
    page_icon="📊",
    layout="wide",
)

# ── API key resolution ────────────────────────────────────────────────────────
def get_api_key() -> str:
    # 1. Environment variable (Streamlit Cloud secret or local setx)
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    # 2. Claude Code local config (for local development)
    claude_config = Path.home() / ".claude" / "config.json"
    if claude_config.exists():
        try:
            with open(claude_config) as f:
                key = json.load(f).get("primaryApiKey", "")
            if key:
                return key
        except Exception:
            pass
    return ""

# ── File reading ──────────────────────────────────────────────────────────────
def read_uploaded_file(name: str, data: bytes) -> str:
    ext = Path(name).suffix.lower()
    try:
        if ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        rows.append("\t".join("" if c is None else str(c) for c in row))
                if rows:
                    parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
            return f"=== {name} ===\n\n" + "\n\n".join(parts)

        elif ext == ".csv":
            import csv
            text = data.decode("utf-8-sig", errors="replace")
            rows = ["\t".join(r) for r in csv.reader(text.splitlines())]
            return f"=== {name} ===\n\n" + "\n".join(rows)

        elif ext == ".docx":
            from docx import Document
            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return f"=== {name} ===\n\n{text}"

        elif ext in (".txt", ".md"):
            return f"=== {name} ===\n\n{data.decode('utf-8', errors='replace')}"

        elif ext == ".json":
            return f"=== {name} ===\n\n{json.dumps(json.loads(data), indent=2)}"

        else:
            return f"[{name}]: Unsupported format ({ext})"
    except Exception as e:
        return f"[{name}]: Error reading file — {e}"


# ── Agent tools ───────────────────────────────────────────────────────────────
def make_tools(file_contents: dict):
    """Return tool implementations bound to the uploaded files."""

    def list_files(_):
        lines = [f"  - {name}" for name in file_contents]
        return "Uploaded files:\n" + "\n".join(lines) if lines else "No files uploaded."

    def read_file(inp):
        name = inp.get("filename", "")
        # fuzzy match
        for key in file_contents:
            if name.lower() in key.lower() or key.lower() in name.lower():
                return file_contents[key]
        available = ", ".join(file_contents.keys())
        return f"File '{name}' not found. Available: {available}"

    def run_python(inp):
        code = inp.get("code", "")
        stdout_buf = io.StringIO()
        try:
            import contextlib
            with contextlib.redirect_stdout(stdout_buf):
                exec(code, {"__builtins__": __builtins__}, {})
            return stdout_buf.getvalue().strip() or "(no output)"
        except Exception:
            return f"ERROR:\n{traceback.format_exc()}"

    tool_map = {
        "list_files": list_files,
        "read_file":  read_file,
        "run_python": run_python,
    }

    tool_defs = [
        {
            "name": "list_files",
            "description": "List all uploaded data room files. Call this first to see what's available.",
            "input_schema": {"type": "object", "properties": {}},
        },
        {
            "name": "read_file",
            "description": "Read the contents of an uploaded file by name.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of the file to read"}
                },
                "required": ["filename"],
            },
        },
        {
            "name": "run_python",
            "description": "Run Python code for calculations. Use print() for output. numpy and basic libs available.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "code": {"type": "string", "description": "Python code to run"}
                },
                "required": ["code"],
            },
        },
    ]

    return tool_defs, tool_map


SYSTEM = """You are an elite M&A analyst. You have been given access to a company's data room files.

Your job:
1. Call list_files first to see what's available
2. Read each relevant file with read_file
3. Use run_python for any financial calculations (CAGRs, margins, retention curves, etc.)
4. Produce a comprehensive due diligence analysis

Structure your final report with these sections:
- Executive Summary
- Revenue & ARR Analysis
- Growth Trends
- Margin Profile
- Customer Concentration
- Churn & Retention
- Key Risks
- Diligence Questions for Management

Always cite specific numbers with file sources. Quantify risks where possible.
Use tables for financial data. Be direct about red flags."""


# ── Agent runner (background thread) ─────────────────────────────────────────
def run_agent_thread(client, messages, tool_defs, tool_map, queue: Queue):
    """Runs the agent loop in a background thread, posting updates to queue."""
    try:
        while True:
            for attempt in range(3):
                try:
                    response = client.messages.create(
                        model="claude-opus-4-6",
                        max_tokens=8096,
                        system=SYSTEM,
                        tools=tool_defs,
                        messages=messages,
                    )
                    break
                except anthropic.RateLimitError:
                    wait = 20 * (attempt + 1)
                    queue.put(("status", f"Rate limit — waiting {wait}s..."))
                    time.sleep(wait)
            else:
                queue.put(("error", "Rate limit exceeded. Please try again in a minute."))
                return

            text_parts = []
            tool_calls = []

            for block in response.content:
                if block.type == "text" and block.text.strip():
                    text_parts.append(block.text)
                elif block.type == "tool_use":
                    tool_calls.append(block)

            if text_parts:
                queue.put(("text", "\n".join(text_parts)))

            if response.stop_reason == "end_turn" or not tool_calls:
                messages.append({"role": "assistant", "content": response.content})
                queue.put(("done", "\n".join(text_parts)))
                return

            messages.append({"role": "assistant", "content": response.content})
            tool_results = []

            for tc in tool_calls:
                args_str = json.dumps(tc.input)
                preview = args_str[:80] + ("..." if len(args_str) > 80 else "")
                queue.put(("tool", f"{tc.name}({preview})"))

                result = tool_map[tc.name](tc.input)
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tc.id,
                    "content": result,
                })

            messages.append({"role": "user", "content": tool_results})

    except Exception as e:
        queue.put(("error", f"Agent error: {e}\n{traceback.format_exc()}"))


# ── UI ────────────────────────────────────────────────────────────────────────
st.title("📊 M&A Analyst Agent")
st.caption("Upload any company's data room files and get a full due diligence analysis.")

# Sidebar
with st.sidebar:
    st.header("How to use")
    st.markdown("""
1. Upload your data room files
2. Click **Run Analysis**
3. Agent reads, calculates & reports
4. Ask follow-up questions below
""")
    st.divider()
    st.markdown("**Supported file types**")
    st.markdown("Excel, CSV, Word, PDF, TXT, JSON")
    st.divider()
    if st.button("Reset / New Deal", use_container_width=True):
        for k in ["messages", "file_contents", "analysis_done", "report"]:
            st.session_state.pop(k, None)
        st.rerun()
    st.caption("Powered by Claude Opus 4.6")

# Initialise state
if "messages"      not in st.session_state: st.session_state.messages      = []
if "file_contents" not in st.session_state: st.session_state.file_contents = {}
if "analysis_done" not in st.session_state: st.session_state.analysis_done = False
if "report"        not in st.session_state: st.session_state.report        = ""

api_key = get_api_key()

# ── Step 1: File upload ──
st.subheader("Step 1 — Upload Data Room Files")
uploaded = st.file_uploader(
    "Drop all files here (Excel, CSV, Word, PDF, TXT)",
    accept_multiple_files=True,
    type=["xlsx", "xls", "csv", "docx", "txt", "md", "json"],
)

if uploaded:
    for f in uploaded:
        if f.name not in st.session_state.file_contents:
            st.session_state.file_contents[f.name] = read_uploaded_file(f.name, f.read())

if st.session_state.file_contents:
    st.success(f"{len(st.session_state.file_contents)} file(s) loaded: {', '.join(st.session_state.file_contents.keys())}")

# ── Step 2: Run analysis ──
st.subheader("Step 2 — Run Analysis")

col1, col2 = st.columns([2, 1])
with col1:
    task = st.text_input(
        "Custom focus (optional)",
        placeholder="e.g. focus on churn risk, or leave blank for full analysis",
    )
with col2:
    run_btn = st.button("Run Analysis", type="primary", use_container_width=True,
                        disabled=not st.session_state.file_contents or not api_key)

if not api_key:
    st.warning("No API key found. Set ANTHROPIC_API_KEY in your environment or Streamlit secrets.")

if run_btn and st.session_state.file_contents:
    # Reset previous analysis
    st.session_state.messages = []
    st.session_state.analysis_done = False
    st.session_state.report = ""

    tool_defs, tool_map = make_tools(st.session_state.file_contents)

    user_task = task.strip() or (
        "Perform a comprehensive M&A due diligence analysis covering: "
        "revenue & ARR trends, growth rates, margins, customer concentration, "
        "churn & retention, key risks, and data gaps."
    )

    messages = [{"role": "user", "content": user_task}]
    st.session_state.messages = messages

    client = anthropic.Anthropic(
        api_key=api_key,
        http_client=httpx.Client(verify=False),
    )

    queue = Queue()
    thread = threading.Thread(
        target=run_agent_thread,
        args=(client, messages, tool_defs, tool_map, queue),
        daemon=True,
    )
    thread.start()

    # Live progress display
    status_box  = st.empty()
    tool_box    = st.empty()
    report_box  = st.empty()

    tool_log = []
    report_text = ""

    status_box.info("Agent is working...")

    while thread.is_alive() or not queue.empty():
        try:
            kind, content = queue.get(timeout=0.3)
        except Empty:
            continue

        if kind == "tool":
            tool_log.append(content)
            tool_box.markdown(
                "**Agent actions:**\n" + "\n".join(f"- `{t}`" for t in tool_log[-10:])
            )
        elif kind == "text":
            report_text += content + "\n"
            report_box.markdown(report_text)
        elif kind == "status":
            status_box.warning(content)
        elif kind == "error":
            status_box.error(content)
            break
        elif kind == "done":
            report_text = content
            st.session_state.report = report_text
            st.session_state.analysis_done = True
            status_box.success("Analysis complete.")
            break

    thread.join(timeout=2)
    st.rerun()

# ── Show completed report ──
if st.session_state.analysis_done and st.session_state.report:
    st.subheader("Analysis Report")
    st.markdown(st.session_state.report)

    st.download_button(
        label="Download Report (.md)",
        data=st.session_state.report,
        file_name="ma_analysis.md",
        mime="text/markdown",
        use_container_width=True,
    )

    # ── Step 3: Follow-up Q&A ──
    st.divider()
    st.subheader("Step 3 — Ask Follow-up Questions")

    # Render chat history (skip first user message which was the task)
    for msg in st.session_state.messages[2:]:
        if isinstance(msg.get("content"), str):
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

    if question := st.chat_input("Ask anything about this deal..."):
        with st.chat_message("user"):
            st.markdown(question)

        st.session_state.messages.append({"role": "user", "content": question})

        tool_defs, tool_map = make_tools(st.session_state.file_contents)
        client = anthropic.Anthropic(
            api_key=api_key,
            http_client=httpx.Client(verify=False),
        )

        queue = Queue()
        thread = threading.Thread(
            target=run_agent_thread,
            args=(client, st.session_state.messages, tool_defs, tool_map, queue),
            daemon=True,
        )
        thread.start()

        with st.chat_message("assistant"):
            placeholder = st.empty()
            answer = ""
            while thread.is_alive() or not queue.empty():
                try:
                    kind, content = queue.get(timeout=0.3)
                except Empty:
                    continue
                if kind in ("text", "done"):
                    answer += content + "\n"
                    placeholder.markdown(answer)
                elif kind == "error":
                    placeholder.error(content)
                    break

            thread.join(timeout=2)

        if answer:
            st.session_state.messages.append({"role": "assistant", "content": answer.strip()})
