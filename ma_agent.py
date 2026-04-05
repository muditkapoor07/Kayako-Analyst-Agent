"""
M&A Analyst Agent — Personal due diligence assistant.
Works on ANY company's data room folder.

Usage:
    python ma_agent.py --deal "path/to/deal/folder"
    python ma_agent.py --deal "path/to/deal/folder" --task "focus on churn risk"
    python ma_agent.py --deals                          # list past deals
    python ma_agent.py --recall "Kayako"                # recall a past deal
"""

import os
import sys
import json
import argparse
import io
import traceback
from contextlib import redirect_stdout, redirect_stderr
from datetime import datetime
from pathlib import Path

import anthropic

# ── Memory directory ──────────────────────────────────────────────────────────
MEMORY_DIR = Path.home() / ".ma_analyst" / "deals"
MEMORY_DIR.mkdir(parents=True, exist_ok=True)

# ── Tool implementations ──────────────────────────────────────────────────────

def list_files(directory: str) -> str:
    """List all files in a directory with sizes and types."""
    p = Path(directory)
    if not p.exists():
        return f"ERROR: Directory '{directory}' does not exist."

    supported = {".xlsx", ".xls", ".csv", ".txt", ".md", ".pdf", ".docx", ".json"}
    files = []
    for f in sorted(p.rglob("*")):
        # skip hidden dirs (.git, __pycache__, etc.) and temp files
        if any(part.startswith(".") or part.startswith("__") for part in f.parts):
            continue
        if f.is_file() and not f.name.startswith("~$") and not f.name.startswith("."):
            size_kb = f.stat().st_size / 1024
            tag = "[readable]" if f.suffix.lower() in supported else "[unsupported]"
            files.append(f"  {f.name:<55} {size_kb:>7.1f} KB  {tag}")

    if not files:
        return f"No files found in '{directory}'."

    return f"Files in '{directory}':\n" + "\n".join(files)


def read_file(filepath: str) -> str:
    """Read contents of a file. Supports Excel, CSV, TXT, MD, DOCX, JSON."""
    p = Path(filepath)
    if not p.exists():
        return f"ERROR: File '{filepath}' not found."

    ext = p.suffix.lower()

    try:
        # ── Excel ──
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        rows.append("\t".join("" if c is None else str(c) for c in row))
                if rows:
                    parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            return f"=== {p.name} ===\n\n" + "\n\n".join(parts)

        # ── CSV ──
        elif ext == ".csv":
            import csv
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = ["\t".join(row) for row in reader]
            return f"=== {p.name} ===\n\n" + "\n".join(rows)

        # ── DOCX ──
        elif ext == ".docx":
            try:
                from docx import Document
                doc = Document(filepath)
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                return f"=== {p.name} ===\n\n{text}"
            except Exception:
                return f"ERROR: Could not read DOCX '{filepath}'. Try python-docx: pip install python-docx"

        # ── PDF ──
        elif ext == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(filepath) as pdf:
                    text = "\n\n".join(page.extract_text() or "" for page in pdf.pages)
                return f"=== {p.name} ===\n\n{text}"
            except Exception:
                return f"ERROR: Could not read PDF. Install pdfplumber: pip install pdfplumber"

        # ── JSON ──
        elif ext == ".json":
            with open(filepath, encoding="utf-8") as f:
                data = json.load(f)
            return f"=== {p.name} ===\n\n{json.dumps(data, indent=2)}"

        # ── Plain text / markdown ──
        elif ext in (".txt", ".md", ".rst"):
            with open(filepath, encoding="utf-8") as f:
                return f"=== {p.name} ===\n\n{f.read()}"

        else:
            return f"Unsupported file type: '{ext}'. Supported: xlsx, csv, docx, pdf, json, txt, md."

    except Exception as e:
        return f"ERROR reading '{filepath}': {e}"


def run_python(code: str) -> str:
    """Execute Python code and return stdout + result. Use for calculations."""
    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()
    local_ns = {}

    try:
        with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
            exec(code, {"__builtins__": __builtins__}, local_ns)
        output = stdout_buf.getvalue()
        errors = stderr_buf.getvalue()
        result = f"{output}"
        if errors:
            result += f"\nSTDERR:\n{errors}"
        return result.strip() or "(no output)"
    except Exception:
        return f"ERROR:\n{traceback.format_exc()}"


def write_output(filename: str, content: str, deal_dir: str = ".") -> str:
    """Write analysis output to a file in the deal folder or outputs/ subfolder."""
    out_dir = Path(deal_dir) / "outputs"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / filename

    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"Saved to: {out_path}"
    except Exception as e:
        return f"ERROR writing file: {e}"


def save_deal_memory(deal_name: str, summary: str) -> str:
    """Save key facts about a deal to persistent memory."""
    mem_file = MEMORY_DIR / f"{deal_name.replace(' ', '_')}.json"
    data = {
        "deal_name":    deal_name,
        "analysed_at":  datetime.now().isoformat(),
        "summary":      summary,
    }
    with open(mem_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    return f"Deal '{deal_name}' saved to memory."


def list_past_deals() -> str:
    """List all previously analysed deals from memory."""
    files = sorted(MEMORY_DIR.glob("*.json"))
    if not files:
        return "No past deals in memory yet."

    lines = []
    for mf in files:
        with open(mf, encoding="utf-8") as f:
            d = json.load(f)
        lines.append(f"  - {d['deal_name']}  (analysed {d['analysed_at'][:10]})")
    return "Past deals:\n" + "\n".join(lines)


def recall_deal(deal_name: str) -> str:
    """Recall summary of a previously analysed deal."""
    # fuzzy match
    files = list(MEMORY_DIR.glob("*.json"))
    for mf in files:
        if deal_name.lower() in mf.stem.lower():
            with open(mf, encoding="utf-8") as f:
                d = json.load(f)
            return f"Deal: {d['deal_name']}\nAnalysed: {d['analysed_at'][:10]}\n\n{d['summary']}"
    return f"No memory found for '{deal_name}'."


# ── Tool registry ─────────────────────────────────────────────────────────────

TOOLS = [
    {
        "name": "list_files",
        "description": (
            "List all files in a directory. ALWAYS call this first when given a new deal folder "
            "to discover what data room files are available before reading anything."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "directory": {"type": "string", "description": "Path to the deal folder"}
            },
            "required": ["directory"],
        },
    },
    {
        "name": "read_file",
        "description": (
            "Read the full contents of a file. Supports Excel (.xlsx/.xls), CSV, DOCX, PDF, "
            "JSON, TXT, and Markdown. Use this to read each data room file."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filepath": {"type": "string", "description": "Full path to the file"}
            },
            "required": ["filepath"],
        },
    },
    {
        "name": "run_python",
        "description": (
            "Execute Python code for financial calculations — CAGRs, margins, ratios, growth rates, "
            "cohort analysis, etc. Always print() your results. "
            "numpy, pandas, and openpyxl are available."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "code": {"type": "string", "description": "Python code to execute. Use print() for output."}
            },
            "required": ["code"],
        },
    },
    {
        "name": "write_output",
        "description": (
            "Save analysis, reports, or summaries to a file inside the deal folder's outputs/ directory. "
            "Use .md for markdown reports, .txt for plain text."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filename":  {"type": "string", "description": "Output filename (e.g. 'risk_analysis.md')"},
                "content":   {"type": "string", "description": "Full content to write"},
                "deal_dir":  {"type": "string", "description": "Deal folder path (for outputs/ subfolder)"},
            },
            "required": ["filename", "content"],
        },
    },
    {
        "name": "save_deal_memory",
        "description": (
            "Save a concise summary of this deal's key facts and metrics to persistent memory. "
            "Call this after completing your initial analysis so you can recall it in future sessions."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "deal_name": {"type": "string", "description": "Short name for this deal (e.g. 'Kayako')"},
                "summary":   {"type": "string", "description": "Key metrics, risks, and observations about this deal"},
            },
            "required": ["deal_name", "summary"],
        },
    },
    {
        "name": "list_past_deals",
        "description": "List all deals previously analysed and stored in memory.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "recall_deal",
        "description": "Recall the saved summary of a previously analysed deal by name.",
        "input_schema": {
            "type": "object",
            "properties": {
                "deal_name": {"type": "string", "description": "Name of the deal to recall"}
            },
            "required": ["deal_name"],
        },
    },
]

TOOL_MAP = {
    "list_files":       lambda i: list_files(i["directory"]),
    "read_file":        lambda i: read_file(i["filepath"]),
    "run_python":       lambda i: run_python(i["code"]),
    "write_output":     lambda i: write_output(i["filename"], i["content"], i.get("deal_dir", ".")),
    "save_deal_memory": lambda i: save_deal_memory(i["deal_name"], i["summary"]),
    "list_past_deals":  lambda i: list_past_deals(),
    "recall_deal":      lambda i: recall_deal(i["deal_name"]),
}

# ── System prompt ─────────────────────────────────────────────────────────────

SYSTEM = """You are an elite M&A analyst with 15 years of experience in SaaS due diligence, \
private equity, and investment banking. You have been hired as a personal analyst assistant.

## Your capabilities
You have tools to:
- Discover and read any files in a deal folder (Excel, CSV, DOCX, PDF, TXT)
- Execute Python for financial calculations (CAGRs, margins, retention curves, etc.)
- Write analysis reports and save them to the deal folder
- Remember deals you've analysed for future reference

## How to approach a new deal
1. ALWAYS start with list_files to see what's in the data room
2. Read the most relevant files (CIM/financials first, then operational data)
3. Use run_python for any non-trivial calculations — don't do maths in your head
4. Identify: ARR/revenue, growth trends, margins, churn/retention, customer concentration, key risks
5. Save your analysis with write_output and save key facts with save_deal_memory

## Analysis standards
- Always cite specific numbers with sources (file name + sheet)
- Flag data gaps clearly — what's missing and why it matters
- Quantify risks where possible (e.g. "$X ARR at risk if top 3 customers churn")
- Maintain professional, investment-committee-ready language
- Be direct about red flags — don't soften material risks

## Output format
Structure responses with clear headings. Use tables for financial data. \
Lead with the most important finding, not preamble."""


# ── Agent loop ────────────────────────────────────────────────────────────────

def run_agent(client: anthropic.Anthropic, messages: list, verbose: bool = True) -> str:
    """Run the agentic loop until Claude stops calling tools. Returns final text."""

    while True:
        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=8096,
            system=SYSTEM,
            tools=TOOLS,
            messages=messages,
        )

        # Collect text from this response
        text_parts = []
        tool_calls = []

        for block in response.content:
            if block.type == "text" and block.text.strip():
                text_parts.append(block.text)
            elif block.type == "tool_use":
                tool_calls.append(block)

        # Print any text Claude produced this turn
        if text_parts and verbose:
            for t in text_parts:
                print(f"\n{t}")

        # Done — no more tool calls
        if response.stop_reason == "end_turn" or not tool_calls:
            final_text = "\n".join(text_parts)
            messages.append({"role": "assistant", "content": response.content})
            return final_text

        # Execute tools
        messages.append({"role": "assistant", "content": response.content})
        tool_results = []

        for tc in tool_calls:
            if verbose:
                args_preview = json.dumps(tc.input)[:120]
                print(f"\n  [tool] {tc.name}({args_preview}{'...' if len(json.dumps(tc.input)) > 120 else ''})")

            result = TOOL_MAP[tc.name](tc.input)

            if verbose and tc.name not in ("save_deal_memory", "list_past_deals"):
                # Show a preview of the result
                preview = result[:300] + "…" if len(result) > 300 else result
                print(f"     → {preview}")

            tool_results.append({
                "type":        "tool_result",
                "tool_use_id": tc.id,
                "content":     result,
            })

        messages.append({"role": "user", "content": tool_results})


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="M&A Analyst Agent — analyse any deal folder with Claude"
    )
    parser.add_argument("--deal",   help="Path to deal data room folder")
    parser.add_argument("--task",   help="Specific analysis task (optional)", default=None)
    parser.add_argument("--deals",  action="store_true", help="List past deals from memory")
    parser.add_argument("--recall", help="Recall a past deal by name", default=None)
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY environment variable.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # ── List past deals ──
    if args.deals:
        print(list_past_deals())
        return

    # ── Recall a deal ──
    if args.recall:
        print(recall_deal(args.recall))
        return

    # ── Analyse a deal ──
    if not args.deal:
        parser.print_help()
        return

    deal_path = Path(args.deal).resolve()
    if not deal_path.exists():
        print(f"ERROR: Deal folder '{deal_path}' does not exist.")
        sys.exit(1)

    task = args.task or (
        "Perform a comprehensive M&A due diligence analysis. "
        "Cover: revenue & ARR trends, growth rates, margins, customer concentration, "
        "churn & retention, key risks, and data gaps. "
        "Generate a written report and save it as 'ma_analysis.md' in the outputs folder. "
        "Then save the deal to memory."
    )

    print(f"\n{'═'*65}")
    print(f"  M&A Analyst Agent")
    print(f"  Deal folder : {deal_path}")
    print(f"  Task        : {task[:80]}{'...' if len(task)>80 else ''}")
    print(f"{'═'*65}\n")

    messages = [{"role": "user", "content": f"Deal folder: {deal_path}\n\nTask: {task}"}]
    run_agent(client, messages)

    # ── Interactive follow-up Q&A ──
    print(f"\n{'─'*65}")
    print("  Analysis complete. Ask follow-up questions (or type 'exit' to quit).")
    print(f"{'─'*65}\n")

    while True:
        try:
            question = input("You: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nGoodbye.")
            break

        if question.lower() in ("exit", "quit", "q", ""):
            print("Goodbye.")
            break

        messages.append({"role": "user", "content": question})
        run_agent(client, messages)


if __name__ == "__main__":
    main()
